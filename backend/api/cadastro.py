import os
import io
from flask import Blueprint, request, jsonify, send_file
import pandas as pd
from backend.core.config import settings
from backend.core.logging import get_logger
from backend.processor import processar_registros_from_files

logger = get_logger()

cadastro_bp = Blueprint('cadastro', __name__, url_prefix='/api')


@cadastro_bp.route('/process_cadastro', methods=['POST'])
def api_process_cadastro():
    try:
        uploaded = request.files.getlist('files[]') or request.files.getlist('files')
        if not uploaded:
            return jsonify({"error": "Nenhum arquivo enviado"}), 400

        paths = []
        for f in uploaded:
            p = os.path.join(settings.UPLOAD_FOLDER, f.filename)
            f.save(p)
            paths.append(p)

        login_choice = request.form.get('login_choice', 'CPF')
        fluxo = request.form.get('fluxo', 'SELF')

        errors, df_final = processar_registros_from_files(paths, login_choice=login_choice, fluxo=fluxo)

        if df_final.empty:
            return jsonify({"error": "Nenhum registro processado", "errors": errors}), 400

        output = io.BytesIO()
        try:
            import openpyxl  # noqa: F401
            from openpyxl.styles import Font, Alignment, PatternFill

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_final.to_excel(writer, sheet_name='Cadastro', index=False)
                ws = writer.sheets['Cadastro']

                header_fill = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
                for cell in list(ws[1]):
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = header_fill

                from openpyxl.utils import get_column_letter
                for idx, col in enumerate(df_final.columns, 1):
                    series = df_final[col].astype(str).fillna("")
                    max_len = max(series.map(len).max(), len(str(col))) + 2
                    max_len = min(max_len, 60)
                    ws.column_dimensions[get_column_letter(idx)].width = max_len

                ws.freeze_panes = 'A2'
                try:
                    ws.auto_filter.ref = ws.dimensions
                except Exception:
                    pass

            output.seek(0)
        except Exception:
            output = io.BytesIO()
            df_final.to_excel(output, index=False)
            output.seek(0)

        for p in paths:
            try:
                if os.path.exists(p):
                    os.remove(p)
            except Exception:
                logger.warning(f"Falha ao remover temporário {p}")

        return send_file(output,
                         download_name="saida_cadastro.xlsx",
                         as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        logger.exception("Erro em /api/process_cadastro")
        return jsonify({"error": str(e)}), 500
