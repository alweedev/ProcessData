import io
import os
import re
import uuid
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
from flask import Blueprint, jsonify, request, send_file

from backend.core.config import settings
from backend.core.logging import get_logger
from backend.utils import format_cpf_for_output, limpar_cpf_raw, upper_no_accents


logger = get_logger()


aprovacao_bp = Blueprint("aprovacao", __name__, url_prefix="/api/aprovacao")


def _normalize_cpf_input(raw_cpf: Optional[str]) -> Tuple[str, str]:
    """Normaliza o CPF de entrada.

    Retorna (cpf_digits, cpf_formatado) ou lança ValueError em caso de CPF inválido.
    """

    if not raw_cpf:
        raise ValueError("Informe um CPF para o aprovador.")

    digits = limpar_cpf_raw(raw_cpf)
    if len(digits) != 11:
        raise ValueError("CPF inválido. Informe 11 dígitos.")

    formatted = format_cpf_for_output(digits)
    return digits, formatted


def _load_users_and_find_approver(users_path: str, cpf_digits: str) -> Tuple[pd.DataFrame, str]:
    """Carrega base de usuários e retorna o nome completo do aprovador.

    Lança ValueError se CPF não existir na base.
    """

    try:
        df_users = pd.read_excel(users_path, dtype=str).fillna("")
    except Exception as exc:  # pragma: no cover - erro de IO
        raise ValueError(f"Falha ao ler base de usuários: {exc}") from exc

    if "CPF" not in df_users.columns:
        raise ValueError("Base de usuários não contém coluna 'CPF'.")

    df_users = df_users.copy()
    df_users["CPFdigits"] = df_users["CPF"].apply(limpar_cpf_raw)
    matches = df_users[df_users["CPFdigits"] == cpf_digits]
    if matches.empty:
        raise ValueError("CPF não encontrado na base de usuários.")

    row = matches.iloc[0]
    nome_completo = str(row.get("NomeCompleto", "")).strip()
    if not nome_completo:
        primeiro = str(row.get("Nome", "")).strip()
        sobrenome = str(row.get("SobreNome", "")).strip()
        nome_completo = f"{primeiro} {sobrenome}".strip()

    return df_users, nome_completo


def _detect_approval_columns(df: pd.DataFrame) -> Dict[str, Any]:
    """Detecta colunas relevantes da base de carga de aprovação.

    Usa nomes esperados, mas de forma case-insensitive.
    Detecta dinamicamente LoginAprovador_1..100.
    """

    col_map: Dict[str, str] = {}
    for col in df.columns:
        key = upper_no_accents(str(col)).replace(" ", "").replace("-", "").replace("_", "")
        col_map[key] = col

    def pick(*candidates: str) -> Optional[str]:
        for cand in candidates:
            key = upper_no_accents(cand).replace(" ", "").replace("-", "").replace("_", "")
            if key in col_map:
                return col_map[key]
        return None

    aprovacao_id = pick("AprovacaoId")
    aprovacao_por = pick("AprovacaoPor")
    aprovacao = pick("Aprovacao")
    tipo = pick("Tipo")
    valor = pick("Valor")
    desc_ccusto = pick("DescricaoCCusto", "DescricaoCentroDeCusto", "DescricaoCCustoEmpresa")
    cod_ccusto = pick("CodigoCCusto", "CodigoCentroDeCusto", "CodigoCCustoEmpresa")
    login_segundo = pick("LoginAprovador_SEGUNDO_NIVEL", "LoginAprovadorSegundoNivel")
    segundo_master = pick("SegundoNivelMaster")
    traveler_name_col = pick("NomeViajante", "NomeCompletoViajante", "NomeCompleto")

    approver_cols: List[str] = []
    for col in df.columns:
        m = re.match(r"(?i)^LoginAprovador_(\d+)$", str(col))
        if m:
            approver_cols.append(col)
    approver_cols.sort(key=lambda c: int(re.search(r"(\d+)$", str(c)).group(1)))

    return {
        "aprovacao_id": aprovacao_id,
        "aprovacao_por": aprovacao_por,
        "aprovacao": aprovacao,
        "tipo": tipo,
        "valor": valor,
        "desc_ccusto": desc_ccusto,
        "cod_ccusto": cod_ccusto,
        "login_segundo": login_segundo,
        "segundo_master": segundo_master,
        "traveler_name_col": traveler_name_col,
        "approver_cols": approver_cols,
    }


def _get_or_create_structure(
    store: Dict[str, Dict[str, Any]],
    row: pd.Series,
    cols: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    """Obtém (ou cria) o registro agregado por AprovacaoId."""

    aprov_id_col = cols.get("aprovacao_id")
    if not aprov_id_col:
        return None

    aprov_id = str(row.get(aprov_id_col, "")).strip()
    if not aprov_id:
        return None

    if aprov_id in store:
        return store[aprov_id]

    aprovacao_por_col = cols.get("aprovacao_por")
    aprovacao_col = cols.get("aprovacao")
    tipo_col = cols.get("tipo")
    valor_col = cols.get("valor")
    desc_ccusto_col = cols.get("desc_ccusto")
    cod_ccusto_col = cols.get("cod_ccusto")
    traveler_col = cols.get("traveler_name_col")

    aprovacao_por_val = str(row.get(aprovacao_por_col, "")).strip() if aprovacao_por_col else ""
    aprovacao_val = str(row.get(aprovacao_col, "")).strip() if aprovacao_col else ""
    tipo_val = str(row.get(tipo_col, "")).strip() if tipo_col else ""
    valor_val = str(row.get(valor_col, "")).strip() if valor_col else ""

    traveler_name_raw = str(row.get(traveler_col, "")).strip() if traveler_col else ""
    cod_cc = str(row.get(cod_ccusto_col, "")).strip() if cod_ccusto_col else ""
    desc_cc = str(row.get(desc_ccusto_col, "")).strip() if desc_ccusto_col else ""
    if cod_cc and desc_cc:
        cost_center = f"{cod_cc} - {desc_cc}"
    else:
        cost_center = cod_cc or desc_cc or ""

    aprov_por_upper = aprovacao_por_val.upper()
    if aprov_por_upper == "VIAJANTE":
        traveler_out: Optional[str] = traveler_name_raw or None
        cost_center_out: Optional[str] = None
    elif aprov_por_upper == "CCEMPRESA":
        traveler_out = None
        cost_center_out = cost_center or None
    else:
        traveler_out = traveler_name_raw or None
        cost_center_out = cost_center or None

    record: Dict[str, Any] = {
        "aprovacao_id": aprov_id,
        "aprovacao_por": aprovacao_por_val or None,
        "aprovacao": aprovacao_val or None,
        "tipo": tipo_val or None,
        "valor": valor_val or None,
        "traveler_name": traveler_out,
        "cost_center": cost_center_out,
        "positions": [],  # preenchido posteriormente
        "in_second_level": False,
        "occurrences_count": 0,
    }

    store[aprov_id] = record
    return record


def _check_structures_without_approvers(
    df_base: pd.DataFrame,
    cpf_digits: str,
    cols: Dict[str, Any],
    target_ids: Set[str],
    remove_second_level: bool,
) -> List[Dict[str, Any]]:
    """Verifica quais estruturas ficarão sem aprovadores após a remoção do CPF.
    
    Retorna lista de estruturas que ficarão vazias (sem nenhum aprovador).
    """
    approver_cols: List[str] = cols.get("approver_cols") or []
    aprov_id_col = cols.get("aprovacao_id")
    login_segundo_col = cols.get("login_segundo")
    
    if not aprov_id_col or not approver_cols:
        return []
    
    structures_without_approvers: List[Dict[str, Any]] = []
    
    for idx, row in df_base.iterrows():
        aprov_id = str(row.get(aprov_id_col, "")).strip()
        if not aprov_id or aprov_id not in target_ids:
            continue
        
        # Contar aprovadores atuais (excluindo o CPF que será removido)
        remaining_approvers: List[str] = []
        for col in approver_cols:
            raw_login = str(row.get(col, "")).strip()
            if not raw_login:
                continue
            # Se for o CPF que será removido, não conta
            if limpar_cpf_raw(raw_login) == cpf_digits:
                continue
            remaining_approvers.append(raw_login)
        
        # Verificar segundo nível (se não estiver sendo removido)
        has_second_level = False
        if login_segundo_col and login_segundo_col in df_base.columns:
            raw_second = str(row.get(login_segundo_col, "")).strip()
            if raw_second:
                # Se remove_second_level=True e o segundo nível é o CPF, não conta
                if remove_second_level and limpar_cpf_raw(raw_second) == cpf_digits:
                    has_second_level = False
                else:
                    has_second_level = True
        
        # Se não sobrar nenhum aprovador, adiciona à lista de alertas
        if len(remaining_approvers) == 0 and not has_second_level:
            aprovacao_por_col = cols.get("aprovacao_por")
            valor_col = cols.get("valor")
            desc_ccusto_col = cols.get("desc_ccusto")
            cod_ccusto_col = cols.get("cod_ccusto")
            traveler_col = cols.get("traveler_name_col")
            
            aprovacao_por_val = str(row.get(aprovacao_por_col, "")).strip() if aprovacao_por_col else ""
            valor_val = str(row.get(valor_col, "")).strip() if valor_col else ""
            
            # Contexto baseado em AprovacaoPor
            contexto = ""
            if aprovacao_por_val.upper() == "VIAJANTE":
                traveler_name = str(row.get(traveler_col, "")).strip() if traveler_col else ""
                contexto = traveler_name or valor_val
            elif aprovacao_por_val.upper() == "CCEMPRESA":
                cod_cc = str(row.get(cod_ccusto_col, "")).strip() if cod_ccusto_col else ""
                desc_cc = str(row.get(desc_ccusto_col, "")).strip() if desc_ccusto_col else ""
                if cod_cc and desc_cc:
                    contexto = f"{cod_cc} - {desc_cc}"
                else:
                    contexto = cod_cc or desc_cc or valor_val
            else:
                contexto = valor_val
            
            structures_without_approvers.append({
                "aprovacaoId": aprov_id,
                "aprovacaoPor": aprovacao_por_val,
                "valor": valor_val,
                "contexto": contexto,
            })
    
    # Remover duplicados por aprovacaoId
    seen: Set[str] = set()
    unique_structures: List[Dict[str, Any]] = []
    for s in structures_without_approvers:
        if s["aprovacaoId"] not in seen:
            seen.add(s["aprovacaoId"])
            unique_structures.append(s)
    
    return unique_structures


def _build_preview_for_cpf(
    df_base: pd.DataFrame,
    cpf_digits: str,
    cols: Dict[str, Any],
    check_empty: bool = False,
    remove_second_level: bool = False,
) -> Dict[str, Any]:
    """Gera estruturas afetadas e estatísticas de preview para um CPF."""

    structures: Dict[str, Dict[str, Any]] = {}

    approver_cols: List[str] = cols.get("approver_cols") or []
    aprov_id_col = cols.get("aprovacao_id")
    if not aprov_id_col or not approver_cols:
        return {"structures": [], "total_structures": 0, "total_occurrences": 0, "affected_ids": [], "structures_without_approvers": []}

    id_vars: List[str] = []
    for key in [
        "aprovacao_id",
        "aprovacao_por",
        "aprovacao",
        "tipo",
        "valor",
        "desc_ccusto",
        "cod_ccusto",
        "traveler_name_col",
        "login_segundo",
    ]:
        colname = cols.get(key)
        if colname and colname in df_base.columns and colname not in id_vars:
            id_vars.append(colname)

    melted = df_base.melt(
        id_vars=id_vars,
        value_vars=approver_cols,
        var_name="slot_col",
        value_name="login",
    ).fillna("")

    melted["CPFdigits"] = melted["login"].apply(limpar_cpf_raw)
    matches_main = melted[melted["CPFdigits"] == cpf_digits]

    for _, row in matches_main.iterrows():
        rec = _get_or_create_structure(structures, row, cols)
        if not rec:
            continue
        slot_col = str(row.get("slot_col", ""))
        m = re.search(r"(\d+)$", slot_col)
        if m:
            pos = int(m.group(1))
            if pos not in rec["positions"]:
                rec["positions"].append(pos)
                rec["occurrences_count"] += 1

    # SEGUNDO_NIVEL
    login_segundo_col = cols.get("login_segundo")
    if login_segundo_col and login_segundo_col in df_base.columns:
        for _, row in df_base.iterrows():
            raw_login = str(row.get(login_segundo_col, "")).strip()
            if not raw_login:
                continue
            if limpar_cpf_raw(raw_login) != cpf_digits:
                continue
            rec = _get_or_create_structure(structures, row, cols)
            if not rec:
                continue
            if not rec["in_second_level"]:
                rec["in_second_level"] = True
            rec["occurrences_count"] += 1

    affected_ids: Set[str] = set(structures.keys())
    total_occurrences = int(sum(rec.get("occurrences_count", 0) for rec in structures.values()))

    ordered_structs = sorted(structures.values(), key=lambda r: r.get("aprovacao_id"))

    # Verificar estruturas que ficarão sem aprovadores
    structures_without_approvers: List[Dict[str, Any]] = []
    if check_empty and affected_ids:
        structures_without_approvers = _check_structures_without_approvers(
            df_base=df_base,
            cpf_digits=cpf_digits,
            cols=cols,
            target_ids=affected_ids,
            remove_second_level=remove_second_level,
        )

    return {
        "structures": ordered_structs,
        "total_structures": len(affected_ids),
        "total_occurrences": total_occurrences,
        "affected_ids": sorted(affected_ids),
        "structures_without_approvers": structures_without_approvers,
    }


def _remove_cpf_and_compact(
    df_base: pd.DataFrame,
    cpf_digits: str,
    cols: Dict[str, Any],
    target_ids: Set[str],
    remove_second_level: bool,
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """Remove todas as ocorrências do CPF e compacta aprovadores 1..100."""

    approver_cols: List[str] = cols.get("approver_cols") or []
    aprov_id_col = cols.get("aprovacao_id")
    if not aprov_id_col or not approver_cols:
        return df_base, {"structures_updated": 0, "occurrences_removed": 0}

    df_out = df_base.copy()
    login_segundo_col = cols.get("login_segundo")
    segundo_master_col = cols.get("segundo_master")

    structures_updated: Set[str] = set()
    occurrences_removed = 0

    for idx, row in df_out.iterrows():
        aprov_id = str(row.get(aprov_id_col, "")).strip()
        if not aprov_id or aprov_id not in target_ids:
            continue

        # Verificar se o CPF aparece nesta linha (main ou segundo nível)
        has_cpf_main = False
        for col in approver_cols:
            raw_login = str(row.get(col, "")).strip()
            if raw_login and limpar_cpf_raw(raw_login) == cpf_digits:
                has_cpf_main = True
                break

        raw_second = str(row.get(login_segundo_col, "")).strip() if login_segundo_col and login_segundo_col in df_out.columns else ""
        has_cpf_second = bool(raw_second and limpar_cpf_raw(raw_second) == cpf_digits)

        if not has_cpf_main and not has_cpf_second:
            continue

        changed = False

        # Remover CPF de LoginAprovador_1..100 e compactar
        if has_cpf_main:
            original_vals: List[str] = [str(row.get(col, "")) for col in approver_cols]
            kept: List[str] = []
            for v in original_vals:
                digits = limpar_cpf_raw(v)
                if digits == cpf_digits and digits:
                    occurrences_removed += 1
                    changed = True
                    continue
                if str(v).strip():
                    kept.append(str(v))

            for pos, col in enumerate(approver_cols):
                new_val = kept[pos] if pos < len(kept) else ""
                df_out.at[idx, col] = new_val

        # Opcionalmente remover do SEGUNDO_NIVEL
        if remove_second_level and has_cpf_second and login_segundo_col and login_segundo_col in df_out.columns:
            df_out.at[idx, login_segundo_col] = ""
            occurrences_removed += 1
            changed = True

        if changed:
            structures_updated.add(aprov_id)

    # Garantir que SegundoNivelMaster permaneça vazio
    if segundo_master_col and segundo_master_col in df_out.columns:
        df_out[segundo_master_col] = df_out[segundo_master_col].astype(str).fillna("")

    stats = {
        "structures_updated": len(structures_updated),
        "occurrences_removed": int(occurrences_removed),
    }
    return df_out, stats


@aprovacao_bp.route("/remover/preview", methods=["POST"])
def aprovacao_remover_preview():
    users_path: Optional[str] = None
    base_path: Optional[str] = None
    try:
        users_file = request.files.get("users_file")
        base_file = request.files.get("base_file")
        form = request.form or {}
        raw_json = request.get_json(silent=True) if request.is_json else None
        cpf_raw = form.get("cpf") or (raw_json or {}).get("cpf")

        # Obter flag de remover segundo nível para cálculo correto
        remove_second_level_raw = form.get("remove_second_level")
        if remove_second_level_raw is None and raw_json is not None:
            remove_second_level_raw = raw_json.get("remove_second_level")
        if isinstance(remove_second_level_raw, bool):
            remove_second_level = remove_second_level_raw
        else:
            remove_second_level = str(remove_second_level_raw or "").lower() in {"1", "true", "yes", "on"}

        if not users_file or not base_file:
            return jsonify({"error": "Envie 'users_file' e 'base_file' (arquivos Excel)."}), 400

        cpf_digits, cpf_formatted = _normalize_cpf_input(cpf_raw)

        # Salvar temporários
        users_ext = os.path.splitext(users_file.filename or "users.xlsx")[1] or ".xlsx"
        base_ext = os.path.splitext(base_file.filename or "base.xlsx")[1] or ".xlsx"

        users_path = os.path.join(settings.UPLOAD_FOLDER, f"{uuid.uuid4().hex}{users_ext}")
        base_path = os.path.join(settings.UPLOAD_FOLDER, f"{uuid.uuid4().hex}{base_ext}")
        users_file.save(users_path)
        base_file.save(base_path)

        _, approver_name = _load_users_and_find_approver(users_path, cpf_digits)

        df_base = pd.read_excel(base_path, dtype=str).fillna("")
        cols = _detect_approval_columns(df_base)

        preview = _build_preview_for_cpf(
            df_base, 
            cpf_digits, 
            cols, 
            check_empty=True, 
            remove_second_level=remove_second_level
        )

        # Converter estruturas internas para o formato esperado pelo frontend
        raw_structures: List[Dict[str, Any]] = preview.get("structures", []) or []
        structures_without_approvers = preview.get("structures_without_approvers", [])
        empty_ids = {s.get("aprovacaoId") for s in structures_without_approvers}

        por_aprovacao_por: Dict[str, int] = {}
        items: List[Dict[str, Any]] = []
        for rec in raw_structures:
            aprovacao_por = (rec.get("aprovacao_por") or "").strip()
            chave_tipo = aprovacao_por or "OUTRO"
            por_aprovacao_por[chave_tipo] = por_aprovacao_por.get(chave_tipo, 0) + 1

            cost_center = rec.get("cost_center") or ""
            cc_codigo: Optional[str] = None
            cc_descricao: Optional[str] = None
            if cost_center:
                # Dividir em "codigo - descricao" se possível
                partes = [p.strip() for p in str(cost_center).split("-", 1)]
                if len(partes) == 2:
                    cc_codigo, cc_descricao = partes[0] or None, partes[1] or None
                else:
                    cc_descricao = partes[0] or None

            positions = rec.get("positions") or []
            try:
                posicoes_norm = [int(p) for p in positions]
            except Exception:
                posicoes_norm = []

            item = {
                "aprovacaoId": rec.get("aprovacao_id"),
                "aprovacaoPor": aprovacao_por or None,
                "aprovacao": rec.get("aprovacao"),
                "tipo": rec.get("tipo"),
                "valor": rec.get("valor"),
                "viajanteNomeCompleto": rec.get("traveler_name"),
                "ccCodigo": cc_codigo,
                "ccDescricao": cc_descricao,
                "posicoes": posicoes_norm,
                "segundoNivel": bool(rec.get("in_second_level")),
                "ficaraSemAprovador": rec.get("aprovacao_id") in empty_ids,
            }
            items.append(item)

        response = {
            "approver": {
                "cpf": cpf_formatted,
                "nomeCompleto": approver_name,
            },
            "summary": {
                "estruturasAfetadas": preview.get("total_structures", 0),
                "ocorrenciasTotal": preview.get("total_occurrences", 0),
                "porAprovacaoPor": por_aprovacao_por,
                "estruturasSemAprovador": len(structures_without_approvers),
            },
            "items": items,
            "alertas": {
                "estruturasSemAprovador": structures_without_approvers,
            },
        }
        return jsonify(response), 200
    except ValueError as ve:
        logger.warning(f"Preview aprovacao remover - erro de validação: {ve}")
        return jsonify({"error": str(ve)}), 400
    except Exception as exc:  # pragma: no cover - proteção extra
        logger.exception("Erro em /api/aprovacao/remover/preview")
        return jsonify({"error": str(exc)}), 500
    finally:
        for path in [users_path, base_path]:
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except Exception as cleanup_exc:  # pragma: no cover
                logger.warning(f"Falha ao remover temporário {path}: {cleanup_exc}")


@aprovacao_bp.route("/remover/export", methods=["POST"])
def aprovacao_remover_export():
    users_path: Optional[str] = None
    base_path: Optional[str] = None
    try:
        users_file = request.files.get("users_file")
        base_file = request.files.get("base_file")
        form = request.form or {}

        raw_json = request.get_json(silent=True) if request.is_json else None

        cpf_raw = form.get("cpf") or (raw_json or {}).get("cpf")
        mode = (form.get("mode") or (raw_json or {}).get("mode") or "all").lower()

        # selected_aprovacao_ids[] pode vir como múltiplos campos de formulário
        selected_ids_form = form.getlist("selected_aprovacao_ids[]") or form.getlist("selected_aprovacao_ids")
        selected_ids_json = (raw_json or {}).get("selected_aprovacao_ids") or []
        selected_ids: Set[str] = set(str(x) for x in (selected_ids_form or selected_ids_json or []))

        remove_second_level_raw = form.get("remove_second_level")
        if remove_second_level_raw is None and raw_json is not None:
            remove_second_level_raw = raw_json.get("remove_second_level")
        if isinstance(remove_second_level_raw, bool):
            remove_second_level = remove_second_level_raw
        else:
            remove_second_level = str(remove_second_level_raw or "").lower() in {"1", "true", "yes", "on"}

        # Flag para ignorar alerta de estruturas vazias
        ignore_empty_warning_raw = form.get("ignore_empty_warning")
        if ignore_empty_warning_raw is None and raw_json is not None:
            ignore_empty_warning_raw = raw_json.get("ignore_empty_warning")
        if isinstance(ignore_empty_warning_raw, bool):
            ignore_empty_warning = ignore_empty_warning_raw
        else:
            ignore_empty_warning = str(ignore_empty_warning_raw or "").lower() in {"1", "true", "yes", "on"}

        if not users_file or not base_file:
            return jsonify({"error": "Envie 'users_file' e 'base_file' (arquivos Excel)."}), 400

        cpf_digits, cpf_formatted = _normalize_cpf_input(cpf_raw)

        # Salvar temporários
        users_ext = os.path.splitext(users_file.filename or "users.xlsx")[1] or ".xlsx"
        base_ext = os.path.splitext(base_file.filename or "base.xlsx")[1] or ".xlsx"

        users_path = os.path.join(settings.UPLOAD_FOLDER, f"{uuid.uuid4().hex}{users_ext}")
        base_path = os.path.join(settings.UPLOAD_FOLDER, f"{uuid.uuid4().hex}{base_ext}")
        users_file.save(users_path)
        base_file.save(base_path)

        # Garante que o CPF existe na base de usuários (e obtém nome apenas para validação/coerência)
        _, _ = _load_users_and_find_approver(users_path, cpf_digits)

        df_base = pd.read_excel(base_path, dtype=str).fillna("")
        cols = _detect_approval_columns(df_base)

        preview = _build_preview_for_cpf(
            df_base, 
            cpf_digits, 
            cols, 
            check_empty=True, 
            remove_second_level=remove_second_level
        )
        affected_ids_all: Set[str] = set(preview.get("affected_ids") or [])
        if not affected_ids_all:
            return jsonify({"error": "CPF não está presente em nenhuma estrutura de aprovação."}), 400

        if mode == "selected":
            if not selected_ids:
                return jsonify({"error": "Informe 'selected_aprovacao_ids' quando mode='selected'."}), 400
            target_ids = affected_ids_all.intersection(selected_ids)
            if not target_ids:
                return jsonify({"error": "Nenhuma AprovacaoId selecionada contém o CPF informado."}), 400
        else:
            target_ids = affected_ids_all

        # Verificar se há estruturas que ficarão sem aprovadores
        structures_without_approvers = preview.get("structures_without_approvers", [])
        affected_empty = [s for s in structures_without_approvers if s.get("aprovacaoId") in target_ids]
        
        if affected_empty and not ignore_empty_warning:
            return jsonify({
                "error": "Algumas estruturas ficarão sem nenhum aprovador após a remoção.",
                "warning": True,
                "estruturasSemAprovador": affected_empty,
                "message": f"{len(affected_empty)} estrutura(s) ficará(ão) sem aprovadores. Deseja continuar mesmo assim?",
            }), 400

        df_updated, stats = _remove_cpf_and_compact(
            df_base=df_base,
            cpf_digits=cpf_digits,
            cols=cols,
            target_ids=target_ids,
            remove_second_level=remove_second_level,
        )

        # Filtrar apenas as estruturas que foram alteradas para reduzir tamanho e tempo
        aprovacao_id_col = cols.get("aprovacao_id")
        if aprovacao_id_col and aprovacao_id_col in df_updated.columns:
            df_export = df_updated[df_updated[aprovacao_id_col].astype(str).isin(target_ids)].copy()
        else:
            df_export = df_updated.copy()

        # Adicionar/atualizar coluna Operacao com valor UPDATE em todas as linhas (como primeira coluna)
        if "Operacao" in df_export.columns:
            df_export["Operacao"] = "UPDATE"
            # Mover para primeira posição se não estiver
            cols_list = df_export.columns.tolist()
            if cols_list[0] != "Operacao":
                cols_list.remove("Operacao")
                cols_list.insert(0, "Operacao")
                df_export = df_export[cols_list]
        else:
            df_export.insert(0, "Operacao", "UPDATE")

        output = io.BytesIO()
        try:
            import openpyxl  # noqa: F401
            from openpyxl.styles import Alignment, Font, PatternFill
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_export.to_excel(writer, sheet_name="Aprovacao", index=False)
                ws = writer.sheets["Aprovacao"]

                header_fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
                for cell in list(ws[1]):
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = header_fill

                from openpyxl.utils import get_column_letter

                for idx, col in enumerate(df_export.columns, 1):
                    series = df_export[col].astype(str).fillna("")
                    max_len = max(series.map(len).max(), len(str(col))) + 2
                    max_len = min(max_len, 60)
                    ws.column_dimensions[get_column_letter(idx)].width = max_len

                ws.freeze_panes = "A2"
                try:
                    ws.auto_filter.ref = ws.dimensions
                except Exception:
                    pass

            output.seek(0)
        except Exception:  # pragma: no cover - fallback simples
            output = io.BytesIO()
            df_export.to_excel(output, index=False)
            output.seek(0)

        filename = f"base_aprovacao_atualizada_{cpf_formatted.replace('-', '')}.xlsx"
        logger.info(
            "Export aprovacao remover gerado - apenas estruturas alteradas",
        )
        logger.info(
            "Estruturas atualizadas: %s | Ocorrencias removidas: %s | Linhas exportadas: %s (de %s total)",
            stats.get("structures_updated"),
            stats.get("occurrences_removed"),
            len(df_export),
            len(df_base),
        )

        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ValueError as ve:
        logger.warning(f"Export aprovacao remover - erro de validação: {ve}")
        return jsonify({"error": str(ve)}), 400
    except Exception as exc:  # pragma: no cover - proteção extra
        logger.exception("Erro em /api/aprovacao/remover/export")
        return jsonify({"error": str(exc)}), 500
    finally:
        for path in [users_path, base_path]:
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except Exception as cleanup_exc:  # pragma: no cover
                logger.warning(f"Falha ao remover temporário {path}: {cleanup_exc}")
