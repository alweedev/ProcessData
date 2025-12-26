import os
import re
import io
import uuid
import pandas as pd
from flask import Blueprint, request, jsonify, send_file

# Use absolute imports to be robust to direct script execution
from backend.core.config import settings
from backend.core.logging import get_logger
from backend.processor import processar_inativacao_from_paths, processar_registros_from_files, MODEL_COLS
from backend.utils import upper_no_accents

logger = get_logger()

inativacao_bp = Blueprint('inativacao', __name__, url_prefix='/api')


def _normalize_lista_columns(df_lista: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza colunas da lista para garantir que 'CPF' e 'NomeCompleto' existam.
    Reutilizável entre preview e process.
    """
    try:
        import re as _re
        norm_map = {_re.sub(r"\s+", " ", str(c)).strip().upper(): c for c in df_lista.columns}

        # CPF: procurar qualquer coluna que contenha 'CPF'
        cpf_src = next((v for k, v in norm_map.items() if 'CPF' in k), None)
        if 'CPF' not in df_lista.columns:
            df_lista['CPF'] = df_lista[cpf_src] if cpf_src else ''

        # NomeCompleto: procurar combinação de Nome + Sobrenome, ou coluna Nome Completo
        nome_src = next((v for k, v in norm_map.items() if 'NOME COMPLETO' in k or 'NOMECOMPLETO' in k), None)
        if 'NomeCompleto' not in df_lista.columns:
            if nome_src:
                df_lista['NomeCompleto'] = df_lista[nome_src]
            else:
                nome = next((v for k, v in norm_map.items() if k == 'NOME' or k.endswith(' NOME')), None)
                sobrenome = next((v for k, v in norm_map.items() if 'SOBRENOME' in k), None)
                if nome and sobrenome:
                    df_lista['NomeCompleto'] = (df_lista[nome].astype(str).fillna('') + ' ' + df_lista[sobrenome].astype(str).fillna('')).str.strip()
                elif nome:
                    df_lista['NomeCompleto'] = df_lista[nome]
                else:
                    any_nome = next((v for k, v in norm_map.items() if 'NOME' in k), None)
                    df_lista['NomeCompleto'] = df_lista[any_nome] if any_nome else ''

        for c in ['CPF', 'NomeCompleto']:
            if c in df_lista.columns:
                df_lista[c] = df_lista[c].astype(str).fillna('')
        # Email: detectar qualquer coluna que contenha 'EMAIL'
        email_src = next((v for k, v in norm_map.items() if 'EMAIL' in k), None)
        if 'Email' not in df_lista.columns:
            df_lista['Email'] = df_lista[email_src] if email_src else ''
        if 'Email' in df_lista.columns:
            df_lista['Email'] = df_lista['Email'].astype(str).fillna('').str.strip()
    except Exception as e:
        logger.warning(f"Normalização de colunas da lista falhou: {e}")

    return df_lista


@inativacao_bp.route("/inativacao/buscar", methods=["POST"])
def api_inativacao_buscar():
    base_path = None
    lista_path = None
    try:
        base_file = request.files.get("base")
        if not base_file:
            return jsonify({"error": "Envie a base (arquivo Excel)"}), 400
        ext = os.path.splitext(base_file.filename)[1]
        base_path = os.path.join(settings.UPLOAD_FOLDER, f"{uuid.uuid4().hex}{ext}")
        base_file.save(base_path)
        df_base = pd.read_excel(base_path, dtype=str).fillna("")

        # Extrair itens (CPFs ou nomes)
        itens = []
        if request.is_json:
            try:
                payload = request.get_json(silent=True) or {}
                itens = payload.get("itens", []) or []
            except Exception:
                itens = []
        if not itens:
            itens_field = request.form.get("itens")
            if itens_field:
                try:
                    import json as _json
                    itens = _json.loads(itens_field)
                except Exception:
                    itens = []
        if not itens:
            lista_file = request.files.get("lista")
            lista_text = request.form.get("lista_text", "")
            if lista_file:
                lista_ext = os.path.splitext(lista_file.filename)[1]
                lista_path = os.path.join(settings.UPLOAD_FOLDER, f"{uuid.uuid4().hex}{lista_ext}")
                lista_file.save(lista_path)
                try:
                    df_lista = pd.read_excel(lista_path, dtype=str).fillna("")
                    df_lista = _normalize_lista_columns(df_lista)
                    if 'CPF' in df_lista.columns:
                        itens = [str(x) for x in df_lista['CPF'].tolist() if str(x).strip()]
                except Exception:
                    itens = []
            elif lista_text.strip():
                itens = [line.strip() for line in lista_text.split('\n') if line.strip()]

        raw_items = [str(x).strip() for x in (itens or [])]
        cpfs_digits = [re.sub(r"\D", "", x) for x in raw_items]
        valid_cpfs = [x for x in cpfs_digits if len(x) == 11]

        # Detectar emails (simples) que não sejam já classificados como CPF
        email_pat = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", re.IGNORECASE)
        valid_emails = [x for x in raw_items if email_pat.match(x) and re.sub(r"\D", "", x) not in valid_cpfs]

        def is_valid_fullname(s: str) -> bool:
            s = upper_no_accents(str(s)).strip()
            parts = s.split()
            return len(parts) >= 2 and len(s) >= 3

        valid_names_raw = [x for x in raw_items if re.sub(r"\D", "", x) not in valid_cpfs and x not in valid_emails and is_valid_fullname(x)]
        valid_names_norm = [upper_no_accents(x).strip() for x in valid_names_raw]

        seen = set()
        duplicates = []
        for c in valid_cpfs:
            if c in seen and c not in duplicates:
                duplicates.append(c)
            seen.add(c)

        def _detect_base_cols(df_base: pd.DataFrame):
            col_map = {upper_no_accents(str(c)).strip(): c for c in df_base.columns}
            cpf_col = next((v for k, v in col_map.items() if "CPF" in k), None)
            nome_col = next((v for k, v in col_map.items() if "NOMECOMPLETO" in k or "NOME COMPLETO" in k or k == 'NOME COMPLETO'), None)
            email_col = next((v for k, v in col_map.items() if "EMAIL" in k), None)
            status_col = next((v for k, v in col_map.items() if "STATUS" in k), None)
            userid_col = next((v for k, v in col_map.items() if "USERID" in k or "IDUSUARIO" in k or k.endswith(' USERID')), None)
            return cpf_col, nome_col, email_col, status_col, userid_col

        cpf_col, nome_col, email_col, status_col, userid_col = _detect_base_cols(df_base)

        df_base = df_base.copy()
        if cpf_col:
            df_base['CPFdigits'] = df_base[cpf_col].apply(lambda v: re.sub(r"\D", "", str(v)))
        else:
            df_base['CPFdigits'] = ""
        if nome_col:
            df_base['NomeNorm'] = df_base[nome_col].apply(lambda v: upper_no_accents(str(v)).strip())
        else:
            df_base['NomeNorm'] = ""

        results = []
        found_cpfs = set()
        found_emails = set()
        if valid_cpfs and cpf_col:
            matches = df_base[df_base['CPFdigits'].isin(valid_cpfs)].copy()
            for _, row in matches.iterrows():
                cpf = row.get('CPFdigits', '')
                found_cpfs.add(cpf)
                item = {
                    "id": str(row.get(userid_col, "")) if userid_col and row.get(userid_col, "") != "" else None,
                    "nome": str(row.get(nome_col, "")) if nome_col else str(row.get('NomeCompleto', "")),
                    "cpf": cpf,
                    "email": str(row.get(email_col, "")) if email_col else str(row.get('Email', "")),
                    "status_atual": str(row.get(status_col, "")) if status_col else str(row.get('Status', "")),
                    "found": True,
                }
                results.append(item)

        found_name_norms = set()
        if valid_names_norm and nome_col:
            name_matches = df_base[df_base['NomeNorm'].isin(valid_names_norm)].copy()
            for _, row in name_matches.iterrows():
                cpf = row.get('CPFdigits', '')
                name_norm = row.get('NomeNorm', '')
                if cpf and cpf in found_cpfs:
                    found_name_norms.add(name_norm)
                    continue
                found_name_norms.add(name_norm)
                item = {
                    "id": str(row.get(userid_col, "")) if userid_col and row.get(userid_col, "") != "" else None,
                    "nome": str(row.get(nome_col, "")) if nome_col else str(row.get('NomeCompleto', "")),
                    "cpf": cpf,
                    "email": str(row.get(email_col, "")) if email_col else str(row.get('Email', "")),
                    "status_atual": str(row.get(status_col, "")) if status_col else str(row.get('Status', "")),
                    "found": True,
                }
                results.append(item)

        # Email matches (case-insensitive)
        if valid_emails and email_col:
            base_email_series = df_base[email_col].astype(str).fillna('')
            base_email_norm = base_email_series.str.strip().str.lower()
            target_emails_norm = [e.strip().lower() for e in valid_emails]
            email_matches = df_base[base_email_norm.isin(target_emails_norm)].copy()
            for _, row in email_matches.iterrows():
                email_val = str(row.get(email_col, '')).strip()
                found_emails.add(email_val.lower())
                item = {
                    "id": str(row.get(userid_col, "")) if userid_col and row.get(userid_col, "") != "" else None,
                    "nome": str(row.get(nome_col, "")) if nome_col else str(row.get('NomeCompleto', "")),
                    "cpf": str(row.get('CPFdigits', '')),
                    "email": email_val,
                    "status_atual": str(row.get(status_col, "")) if status_col else str(row.get('Status', "")),
                    "found": True,
                }
                results.append(item)

        not_found_cpfs = [c for c in valid_cpfs if c not in found_cpfs]
        for cpf in not_found_cpfs:
            results.append({
                "id": None,
                "nome": "",
                "cpf": cpf,
                "email": "",
                "status_atual": "Não localizado",
                "found": False,
            })
        not_found_names = [n for n in valid_names_norm if n not in found_name_norms]
        for n in not_found_names:
            results.append({
                "id": None,
                "nome": n,
                "cpf": "",
                "email": "",
                "status_atual": "Não localizado",
                "found": False,
            })
        not_found_emails = [e for e in valid_emails if e.strip().lower() not in found_emails]
        for e in not_found_emails:
            results.append({
                "id": None,
                "nome": "",
                "cpf": "",
                "email": e,
                "status_atual": "Não localizado",
                "found": False,
            })

        def sort_key(it):
            return (0 if it.get('found') else 1, it.get('nome') or '', it.get('cpf') or '')
        results = sorted(results, key=sort_key)

        return jsonify({
            "items": results,
            "total": len(results),
            "duplicates": duplicates,  # pode conter CPFs duplicados; emails duplicados não são listados separadamente
            "not_found": not_found_cpfs + valid_names_raw + not_found_emails,
        }), 200
    except Exception as e:
        logger.exception("Erro em /api/inativacao/buscar")
        return jsonify({"error": str(e)}), 500
    finally:
        for path in [base_path, lista_path]:
            try:
                if path and os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass


@inativacao_bp.route("/inativacao/executar", methods=["POST"])
def api_inativacao_executar():
    try:
        if not request.is_json:
            return jsonify({"error": "Conteúdo inválido. Envie JSON."}), 400
        data = request.get_json(silent=True) or {}
        usuarios = data.get("usuarios", [])
        if not usuarios:
            return jsonify({"error": "Nenhum usuário selecionado"}), 400
        processed = []
        seen = set()
        for u in usuarios:
            cpf = re.sub(r"\D", "", str(u.get('cpf', '')))
            if len(cpf) != 11:
                continue
            if cpf in seen:
                continue
            seen.add(cpf)
            processed.append({"id": u.get('id'), "cpf": cpf})
        return jsonify({
            "success": True,
            "processed": len(processed),
            "usuarios": processed,
            "message": f"{len(processed)} usuário(s) inativados."
        }), 200
    except Exception as e:
        logger.exception("Erro em /api/inativacao/executar")
        return jsonify({"error": str(e)}), 500


@inativacao_bp.route("/process_inativacao", methods=["POST"])
def api_process_inativacao():
    base_path = None
    lista_path = None
    try:
        base_file = request.files.get("base")
        lista_file = request.files.get("lista")
        lista_text = request.form.get("lista_text", "").strip()

        if not base_file:
            logger.error("Nenhum arquivo 'base' enviado")
            return jsonify({"error": "Envie a base"}), 400

        if not (lista_file or lista_text):
            logger.error("Nenhum arquivo 'lista' ou texto enviado")
            return jsonify({"error": "Envie a lista ou insira os nomes/CPFs"}), 400

        ext = os.path.splitext(base_file.filename)[1]
        base_path = os.path.join(settings.UPLOAD_FOLDER, f"{uuid.uuid4().hex}{ext}")
        base_file.save(base_path)
        logger.info(f"Arquivo base salvo em: {base_path}")

        if lista_file:
            lista_ext = os.path.splitext(lista_file.filename)[1]
            lista_path = os.path.join(settings.UPLOAD_FOLDER, f"{uuid.uuid4().hex}{lista_ext}")
            lista_file.save(lista_path)
            logger.info(f"Arquivo lista salvo em: {lista_path}")
            df_lista = pd.read_excel(lista_path, dtype=str).fillna("")
            df_lista = _normalize_lista_columns(df_lista)
        else:
            logger.info("Processando lista a partir de texto")
            lista_items = [item.strip() for item in lista_text.split('\n') if item.strip()]
            if not lista_items:
                return jsonify({"error": "Texto de lista vazio ou sem CPF/Nome/E-mail válidos"}), 400
            email_pat = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", re.IGNORECASE)
            rows = []
            for it in lista_items:
                digits = re.sub(r"\D", "", it)
                if email_pat.match(it):
                    rows.append({"CPF": "", "NomeCompleto": "", "Email": it})
                elif len(digits) == 11:
                    rows.append({"CPF": it, "NomeCompleto": "", "Email": ""})
                else:
                    rows.append({"CPF": "", "NomeCompleto": it, "Email": ""})
            df_lista = pd.DataFrame(rows)
            if 'Email' not in df_lista.columns:
                df_lista['Email'] = ''

        df_base = pd.read_excel(base_path, dtype=str).fillna("")

        use_fuzzy = request.form.get('use_fuzzy', 'false').lower() in ['1', 'true', 'yes']
        try:
            fuzzy_cutoff = float(request.form.get('fuzzy_cutoff', 0.90))
        except Exception:
            fuzzy_cutoff = 0.90

        out = processar_inativacao_from_paths(df_base, df_lista, use_fuzzy=use_fuzzy, fuzzy_cutoff=fuzzy_cutoff)
        if isinstance(out, tuple) and len(out) == 2:
            out_df, stats = out
        else:
            out_df = out
            stats = {}

        logger.info(f"DataFrame gerado: {out_df.shape} linhas, {out_df.columns.tolist()} colunas")
        if out_df.empty:
            logger.warning("DataFrame vazio retornado por processar_inativacao_from_paths")
            if stats and stats.get('inactive_matches'):
                return jsonify({"error": "Nenhuma linha ativa correspondeu; foram encontradas correspondências INATIVAS.", "stats": stats}), 400
            return jsonify({"error": "Nenhum dado processado para inativação", "stats": stats}), 400

        output = io.BytesIO()
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                out_df.to_excel(writer, sheet_name='Inativacao', index=False)
                ws = writer.sheets['Inativacao']
                header_fill = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
                for cell in list(ws[1]):
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = header_fill
                from openpyxl.utils import get_column_letter
                for idx, col in enumerate(out_df.columns, 1):
                    series = out_df[col].astype(str).fillna("")
                    max_len = max(series.map(len).max(), len(str(col))) + 2
                    max_len = min(max_len, 60)
                    ws.column_dimensions[get_column_letter(idx)].width = max_len
                ws.freeze_panes = 'A2'
                try:
                    ws.auto_filter.ref = ws.dimensions
                except Exception:
                    pass
            output.seek(0)
        except Exception:
            output = io.BytesIO()
            out_df.to_excel(output, index=False)
            output.seek(0)

        logger.info("Arquivo de inativação gerado e enviado")
        return send_file(output,
                         download_name="saida_inativacao.xlsx",
                         as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        logger.exception("Erro em /api/process_inativacao")
        return jsonify({"error": str(e)}), 500
    finally:
        for path in [base_path, lista_path]:
            try:
                if path and os.path.exists(path):
                    os.remove(path)
                    logger.info(f"Arquivo temporário removido: {path}")
            except Exception as e:
                logger.warning(f"Falha ao remover {path}: {str(e)}")


@inativacao_bp.route("/preview_inativacao", methods=["POST"])
def api_preview_inativacao():
    base_path = None
    lista_path = None
    try:
        base_file = request.files.get("base")
        lista_file = request.files.get("lista")
        lista_text = request.form.get("lista_text", "").strip()

        if not base_file:
            return jsonify({"error": "Envie a base"}), 400

        ext = os.path.splitext(base_file.filename)[1]
        base_path = os.path.join(settings.UPLOAD_FOLDER, f"{uuid.uuid4().hex}{ext}")
        base_file.save(base_path)

        if lista_file:
            lista_ext = os.path.splitext(lista_file.filename)[1]
            lista_path = os.path.join(settings.UPLOAD_FOLDER, f"{uuid.uuid4().hex}{lista_ext}")
            lista_file.save(lista_path)
            df_lista = pd.read_excel(lista_path, dtype=str).fillna("")
            df_lista = _normalize_lista_columns(df_lista)
        else:
            if not lista_text:
                return jsonify({"error": "Envie a lista como arquivo ou cole nomes/CPFs no campo de texto"}), 400
            lista_items = [item.strip() for item in lista_text.split('\n') if item.strip()]
            if not lista_items:
                return jsonify({"error": "Texto de lista vazio ou sem CPF/Nome/E-mail válidos"}), 400
            email_pat = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", re.IGNORECASE)
            rows = []
            for it in lista_items:
                digits = re.sub(r"\D", "", it)
                if email_pat.match(it):
                    rows.append({"CPF": "", "NomeCompleto": "", "Email": it})
                elif len(digits) == 11:
                    rows.append({"CPF": it, "NomeCompleto": "", "Email": ""})
                else:
                    rows.append({"CPF": "", "NomeCompleto": it, "Email": ""})
            df_lista = pd.DataFrame(rows)
            if 'Email' not in df_lista.columns:
                df_lista['Email'] = ''

        df_base = pd.read_excel(base_path, dtype=str).fillna("")

        use_fuzzy = request.form.get('use_fuzzy', 'false').lower() in ['1', 'true', 'yes']
        try:
            fuzzy_cutoff = float(request.form.get('fuzzy_cutoff', 0.90))
        except Exception:
            fuzzy_cutoff = 0.90
        out = processar_inativacao_from_paths(df_base, df_lista, use_fuzzy=use_fuzzy, fuzzy_cutoff=fuzzy_cutoff)
        if isinstance(out, tuple) and len(out) == 2:
            out_df, stats = out
        else:
            out_df = out
            stats = {}

        try:
            count = int(stats.get('total_matches')) if stats and 'total_matches' in stats else (int(out_df.shape[0]) if out_df is not None else 0)
        except Exception:
            count = int(out_df.shape[0]) if out_df is not None else 0

        sample = []
        columns = []
        records = []
        if out_df is not None and not out_df.empty:
            try:
                columns = list(out_df.columns)
            except Exception:
                columns = []
            try:
                sample = out_df.head(10).to_dict(orient="records")
            except Exception:
                sample = []
            try:
                records = out_df.head(500).to_dict(orient='records')
            except Exception:
                records = sample[:]
        else:
            try:
                inactive = (stats or {}).get('inactive_matches') or {}
                all_rows = []
                for k, lst in (inactive.items() if isinstance(inactive, dict) else []):
                    if isinstance(lst, list):
                        for it in lst:
                            try:
                                r = dict(it)
                                r['match_type'] = k
                                all_rows.append(r)
                            except Exception:
                                pass
                if all_rows:
                    colset = set()
                    for r in all_rows[:50]:
                        try:
                            colset.update(list(r.keys()))
                        except Exception:
                            pass
                    columns = list(colset)
                    sample = all_rows[:10]
                    records = all_rows[:500]
            except Exception:
                pass

        try:
            logger.info(f"/api/preview_inativacao -> count={count} sample={len(sample)} records={len(records)} columns={len(columns)}")
            if sample and isinstance(sample, list) and len(sample) > 0:
                logger.debug(f"preview sample keys: {list(sample[0].keys())[:10]}")
        except Exception:
            pass
        return jsonify({"count": count, "sample": sample, "columns": columns, "records": records, "stats": stats}), 200
    except Exception as e:
        logger.exception("Erro em /api/preview_inativacao")
        return jsonify({"error": str(e)}), 500
    finally:
        for path in [base_path, lista_path]:
            try:
                if path and os.path.exists(path):
                    os.remove(path)
                    logger.info(f"Arquivo temporário removido: {path}")
            except Exception as e:
                logger.warning(f"Falha ao remover {path}: {str(e)}")
